# Copyright (c) 2016 Joe Vernaci
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import collections
import csv
import functools
import logging
import math
import os
import sys
import unicodedata

try:
    from openpyxl import Workbook
    from openpyxl import __version__ as openpyxl_version
    from openpyxl.cell.cell import KNOWN_TYPES
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.drawing.image import _import_image
    from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    pass  # openpyxl not required.

try:
    from PIL import Image as PILImage
except ImportError:
    pass


from . import prefs
from . import util
from .exceptions import PreferencesTypeError, ExportError


log = logging.getLogger(__name__)


FMT_CURRENCY = 0


# The following constants and functions are needed by preferences.  Since
# they are added ad the top level they must go before the prefs.add.
_XLSXCurrencyFormatMap = {
    'USD': '$#,##0.00;-$#,##0.00',
    'GBP': u'[${}-809]#,##0.00;-[${}-809]#,##0.00'.format(
            unicodedata.lookup('POUND SIGN'),
            unicodedata.lookup('POUND SIGN')),
    'EUR': u'[${}-1809]#,##0.00;[RED]-[${}-1809]#,##0.00'.format(
            unicodedata.lookup('EURO SIGN'),
            unicodedata.lookup('EURO SIGN')),
}


def _pref_xlsx_currency_fmt_type(value):
    """type converter for export.xlsx.currency_fmt preference.

    Args:
        value: string to convert.

    Returns:
        matching string.
    """
    value = str(value).upper()
    keys = _XLSXCurrencyFormatMap.keys()
    if value not in keys:
        log.info('currency_fmt must be in {}'.format(keys))
        raise PreferencesTypeError('currency_fmt must be in {}'.format(keys))
    return value


def _mp_write_rels(self):
    """Replacement function for
    openpyxl.drawing.spreadsheet_drawing.SpreadsheetDrawing._write_rels.
    _write_rels generates a lxml.etree object to create the xlsx relationship
    xml file that maps embedded images to actual files.  _write_rels hardcodes
    the extensions to .png.  This file will replace the extension with what is
    stored in the path stored in the pillow image object.

    Returns:
        modified lxml.etree.
    """
    tree = self.__mp_write_rels()
    if not len(self.images):
        return tree

    if hasattr(self.images[0], 'path'):
        # Get path version >= 2.4.0
        image_name_mapping = [os.path.basename(x.path) for x in self.images]
    else:
        # Get path version < 2.4.0
        image_name_mapping = [os.path.basename(x._path) for x in self.images]
    image_name_mapping = {os.path.splitext(x)[0]: x for x
            in image_name_mapping}

    for rel in tree.iterchildren():
        rel_dir, rel_file = os.path.split(rel.get('Target'))
        img_file = image_name_mapping.get(os.path.splitext(
                rel_file)[0], None)

        if img_file is None or img_file == rel_file:
            continue

        log.debug('changing rel {} -> {}'.format(rel_file, img_file))
        rel.set('Target', os.path.join(rel_dir, img_file))
    return tree


def is_mp_write_rels():
    """Returns tree if _write_rels has been monkey-patched with
    _mp_write_rels."""
    return SpreadsheetDrawing._write_rels.im_func == _mp_write_rels


def mp_write_rels():
    """Monkey-patches _write_rels with _mp_write_rels."""
    if is_mp_write_rels() is True:
        return

    SpreadsheetDrawing.__mp_write_rels = SpreadsheetDrawing._write_rels
    SpreadsheetDrawing._write_rels = _mp_write_rels


def mr_write_rels():
    """Reverts moneky-patch of _write_rels with _mp_write_rels."""
    if is_mp_write_rels() is False:
        return

    SpreadsheetDrawing._write_rels = SpreadsheetDrawing.__mp_write_rels
    del SpreadsheetDrawing.__mp_write_rels


def _pref_jpeg_en_on_change(pref, value, user_data):
    """Callback for export.xlsx.img.jpeg.enable preference change.  This
    is used to trigger monkey-patching/reverting _write_rels.

    Args:
        pref: calling preference
        value: bool
        user_data: unused
    """
    log.debug('value jpeg_en changed to: {}'.format(value))
    try:
        if value is True:
            mp_write_rels()
        else:
            mr_write_rels()
    except NameError:
        pass


prefs.add('export.default_csv', bool, True, default=False,
        help='True to export using csv files')
prefs.add('export.image.resize.max_height', int, True, default=200,
        help='Max height in pixels for image exports')
prefs.add('export.image.resize.max_width', int, True, default=200,
        help='Max width in pixels for image exports')
prefs.add('export.tree.use_cache', bool, True, default=True,
        help='True to export using cache values from tree view.  False to '
        'always export fresh objects from the database')
prefs.add('export.xlsx.currency_fmt', _pref_xlsx_currency_fmt_type, True,
        default='USD', help='Currency format used in xlsx exports.  Valid '
        'choices are (USD,GBP,EUR)')
prefs.add('export.xlsx.img.jpeg.enable', bool, True, default=True,
        on_change=_pref_jpeg_en_on_change,
        help='(EXPERIMENTAL) Convert images to jpeg when exporting xlsx.  '
        'Openpyxl only uses png and this requires monkey patching openpyxl '
        'and pillow')
prefs.add('export.xlsx.img.jpeg.quality', int, True, default=90,
        help='quality setting for pillow when exporting jpeg images')
prefs.add('export.xlsx.img.multiplier.height', float, True, default=0.76,
        help='xlsx row ht attribute multiplier for images (per pixels)')
prefs.add('export.xlsx.img.multiplier.width', float, True, default=0.108,
        help='xlsx col width attribute multiplier for images (per pixels)')
prefs.add('export.xlsx.text.multiplier.width', float, True, default=1.0,
        help='xlsx col width attribute multiplier text (per char)')


def key_intersection(export_keys, obj_keys, obj_type, warn_on_missing=True):
    """Performs set intersection of export and object keys.

    Args:
        export_keys: list of keys to used for export.
        obj_keys: list of keys in object to export.
        obj_type: string name of object (used for logging).
        warn_on_missing (optional): True to log warning if export keys are not
            found in object keys.

    Returns:
        list of intersected keys.
    """
    if export_keys is None:
        return obj_keys

    obj_keys = set(obj_keys)
    exp_keys = set(export_keys)

    missing_keys = exp_keys - obj_keys
    if warn_on_missing is True and len(missing_keys):
        missing_keys = ', '.join(missing_keys)
        err = '{} keys missing from {}'.format(missing_keys, obj_type)
        log.warn(err)

    # Use list comprehension (not just set.__and__) to maintain key
    # order from export_keys.
    obj_keys = obj_keys & exp_keys
    obj_keys = [x for x in export_keys if x in obj_keys]
    return obj_keys


class CSVExporter(object):
    default_ext = 'csv'
    delimiter = ','
    quotechar = '"'
    img_support = False

    def __init__(self, path):
        """Initializer for CSV exporter.

        Args:
            path: path string for csv file to write.
        """
        self.file = open(path, 'w')
        self.writer = csv.writer(self.file, delimiter=self.delimiter,
                quotechar=self.quotechar, quoting=csv.QUOTE_MINIMAL)

    def close(self):
        """Close export file."""
        self.file.close()
        self.file = None
        self.writer = None

    def write_row(self, iterable):
        """Writes single row to csv file.

        Args:
            iterable: string or list of strings.  One item per column.
        """
        if isinstance(iterable, basestring):
            iterable = [iterable,]
        self.writer.writerow(iterable)

    def write_hdr(self, iterable):
        """Writes single header row to csv file (used for compatibility with
        other exporters).

        Args:
            iterable: string or list of strings.  One item per column.
        """
        self.write_row(iterable)

    def write_blank(self, count=1):
        """Write blank lines to csv file.

        Args:
            count (optional): int number of blank lines to write.
        """
        for x in range(count):
            self.write_row(list())

    def image(self, *args, **kwargs):
        """Returns warning string that image is not supported (used for
        compatibility with other exporters).

        Args:
            *args: ignored
            **kwargs: ignored
        """
        return 'images not supported in CSV'


# The support check functions are done during runtime to wait for preferences
# to be loaded so they do not log warnings if unused.
__openpyxl_support = None
def _openpyxl_support(simulate_no_openpyxl=False):
    """Checks once if openpyxl is supported.  Called during runtime to only
    warn if attempting to use openpyxl when not supported.

    Args:
        simulate_no_openpyxl (optional): bool to simulate openpyxl not
        installed (used for development).
    Returns:
        True if supported, False if not.
    """
    global __openpyxl_support
    if __openpyxl_support is None:
        if simulate_no_openpyxl is False \
                and sys.modules.has_key('openpyxl') is True:
            __openpyxl_support = True
            log.debug('using openpyxl version: {}'.format(openpyxl_version))
        else:
            __openpyxl_support = False
            log.warn('openpyxl not installed, falling back to CSVExporter')

    return __openpyxl_support


__openpyxl_img_support = None
def _openpyxl_img_support(simulate_no_pil=False):
    """Checks once if openpyxl images are supported (PIL/pillow is installed).
    Called during runtime to only warn if attempting to use openpyxl images
    when not supported.

    Args:
        simulate_no_pil (optional): bool to simulate PIL/pillow not
        installed (used for development).
    Returns:
        True if supported, False if not.
    """
    global __openpyxl_img_support
    if __openpyxl_img_support is None:
        if simulate_no_pil is True:
            original_sys_path = sys.path[:]
            sys.path.remove('/usr/lib/python2.7/dist-packages')

        try:
            _import_image(None)
        except ImportError:
            log.warn('PIL not installed, exporting images disabled')
            __openpyxl_img_support = False
        except AttributeError:
            __openpyxl_img_support = True

        if simulate_no_pil is True:
            sys.path = original_sys_path

    return __openpyxl_img_support


# Use JPEG for format in PIL
try:
    class OpenpyxlImageExt(OpenpyxlImage):
        """Subclass of openpyxl image that allows changing image file
        extension in the object._path attribute.  In openpyxl 2.3.5 and
        below this attribute is read only.
        """
        _ext_value = None

        @property
        def _ext(self):
            return self._ext_value

        @_ext.setter
        def _ext(self, value):
            if value == '':
                value = None
            elif not value.startswith(os.path.extsep):
                value = os.path.extsep + value
            self._ext_value = value

        @property
        def _path(self):
            path = super(OpenpyxlImageExt, self)._path
            if self._ext is not None:
                path = os.path.splitext(path)[0] + self._ext
            return path

except NameError:
    pass


def _mp_pil_image_save(self, _mp_func, _mp_fmt, _mp_params, fp, format=None,
        **params):
    """Replacement function for PIL.Image.Image.save.  This does not change
    any functionality of the function but allows for overrides to be put in
    place.  This works around openpyxl.writer.excel.ExcelWriter._write_images
    which hardcodes image format to .png.  It also allows adding extra
    parameters such as quality.  This should be used as a base for a
    functools.partial function to expose only the same arguments to the
    rest of the program.

    Args:
        _mp_func: original function.
        _mp_fmt: format string to override with.
        _mp_params: dict of keyword arguments to override in params.
        fp: file to save to.
        format: ignored.
        **params (optional): keyword arguments passed to the image writer.
    """
    params.update(_mp_params)
    ret = _mp_func(fp, format=_mp_fmt, **params)
    # Reset file position for openpyxl 2.4.5 or it will write 0 sized images.
    if hasattr(fp, 'seek'):
        fp.seek(0)
    return ret


def is_mp_pil_image_save(img):
    """Returns tree if img.save has been monkey-patched with
    a partial function of _mp_pil_image_save."""
    if not isinstance(img, PILImage.Image):
        raise TypeError('img must be type PIL.Image.Image')

    if not isinstance(img.save, functools.partial):
        return False

    return img.save.func.im_func == _mp_pil_image_save


def mp_pil_image_save(img, fmt, params):
    """Monkey-patches img with a partial function of _mp_pil_image_save.

    Args:
        img: PIL.Image.Image object.
        fmt: format string to override with.
        params: dict of keyword arguments to override in params.

    Returns:
        patched img.
    """
    if is_mp_pil_image_save is True:
        raise ValueError('img already monkey patched')

    func = _mp_pil_image_save.__get__(img, img.__class__)
    func = functools.partial(func, img.save, fmt, params)
    img.save = func

    return img


def mr_pil_image_save(img):
    """Reverts monkey-patched of img with a partial function of
    _mp_pil_image_save.

    Args:
        img: patched PIL.Image.Image object.

    Returns:
        img.
    """
    if is_mp_pil_image_save is True:
        img.save = img.save.args[0]
    return img


class XLSXCurrency(float):
    """Provides unique type and __len__ for write_row.  This is a subclass of
    float instead of string so setting number_format in the cell works.
    """
    def __len__(self):
        # More conservative to use round instead of int to compensate for
        # inaccuracy in log (math.log(1000, 10) < 3).
        length = int(round(math.log(self, 10)))
        # Add commas.
        length += length / 3
        # ones place (1) + decimal/cents (3) + symbols (1) + negative (1)
        length += 6
        return length


class XLSXExporter(object):
    # Notes on column width:
    #
    # From:
    # https://msdn.microsoft.com/en-us/library/documentformat.openxml.spreadsheet.column.aspx
    # Column width measured as the number of characters of the maximum
    # digit width of the numbers 0, 1, 2, ..., 9 as rendered in the
    # normal style's font.
    # ...
    # width = Truncate([{Number of Characters} * {Maximum Digit Width}
    #       + {5 pixel padding}] / {Maximum Digit Width} * 256) / 256
    # Where {Maximum Digit Width} is in pixels.  The exact width can not
    # be calculated without this value.  Doing a 1 to 1 mapping of
    # {Number of Characters} to width seems to work well enough.  Though
    # this is going to be dependent on the users default font.  The user
    # can set export.xlsx.text.multiplier.width to modify the width
    # calculation.
    #
    # There is also the bestFit attribute in Column properties
    # (set by column_dimensions[index].auto_size = True) but it does not
    # appear to be used by Calc or Excel.
    #
    # Notes on row height:
    # From:
    # https://msdn.microsoft.com/en-us/library/documentformat.openxml.spreadsheet.row.aspx
    # Row height measured in point size. There is no margin padding on
    # row height.
    #
    # This again seems to be dependent on what the default font size for
    # the application will be.  For text only rows we can leave this
    # unset (it will be unset in the XML as well) and Calc will pick
    # an appropriate default.
    #
    # Image multipliers are just found empirically by inserting an image
    # of know size, adjusting the row height and column width
    # appropriately and examining the xml.
    default_ext = 'xlsx'

    def __init__(self, path):
        """Initializer for XLSX exporter.

        Args:
            path: path string for csv file to write.
        """
        self.path = path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.widths = []
        self._other_widths = []
        self.img_support = _openpyxl_img_support()

    def close(self):
        """Close export file."""
        self._set_col_widths()

        self.wb.save(self.path)
        self.ws = None
        self.wb = None
        log.debug('xlsx column widths: {}'.format(self.widths))

    def _set_col_widths(self):
        """Sets column width attributes in export file with values saved
        in self.widths."""
        for index in range(self.ws.max_column):
            width = util.getindex(self.widths, index, default=None)
            if width is None:
                continue

            # get_column_letter indexing starts at 1.
            index = get_column_letter(index + 1)
            self.ws.column_dimensions[index].width = width

    def _apply_width(self, widths):
        """Sets max of self.widths and new row widths.

        Args:
            widths: list of floats of column widths for current row.
        """
        if len(self.widths) < len(widths):
            self.widths.extend([None,] * (len(widths) - len(self.widths)))

        for index, current, new in zip(range(len(self.widths)),
                self.widths, widths):
            if current < new:
                self.widths[index] = new

    def _apply_row_number_format(self, row, col_fmt_pairs):
        """Sets number_format in cells for a given row.

        Args:
            row: int row index.
            col_fmt_pairs: list of tuples of (column index, number_format).
        """
        for col, fmt in col_fmt_pairs:
            if fmt == FMT_CURRENCY:
                fmt = _XLSXCurrencyFormatMap[
                        prefs['export.xlsx.currency_fmt']]
            else:
                raise ValueError('unknown fmt: \'{}\''.format(fmt))

            cell = self.ws.cell(column=col, row=row)
            cell.number_format = fmt

    def _apply_row_height(self, row, height):
        """Sets row height.

        Args:
            row: int row index.
            height: float of row height.
        """
        if height is not None:
            self.ws.row_dimensions[row].height = height

    def _apply_img(self, row, col_img_pairs):
        """Embeds images incells for a given row.

        Args:
            row: int row index.
            col_img_pairs: list of tuples of (column index, openpyxl image).
        """
        for col, img in col_img_pairs:
            if not isinstance(img, OpenpyxlImage):
                raise TypeError('img must be openpyxl Image')

            cell = self.ws.cell(column=col, row=row)
            self.ws.add_image(img, cell.coordinate)

    def write_row(self, iterable, multiplier=1):
        """Writes single row to xlsx file.

        Args:
            iterable: string or list openpyxl.cell.cell.KNOWN_TYPES or
                openpyxl/PIL images.
            multiplier (optional): float multiplier applied to calculated
                column width.
        """
        if isinstance(iterable, basestring) \
                or not isinstance(iterable, collections.Iterable):
            iterable = [iterable,]

        text_multiplier = prefs['export.xlsx.text.multiplier.width'] \
                * multiplier

        # Dense data, done for every row.
        cell_text = []
        row_widths = []

        # Sparse data, done on a per column basis.
        col_img_pairs = []
        col_fmt_pairs = []

        row_height = None

        # openpyxl column indexing starts at 1 for A.
        for col, value in enumerate(iterable, 1):
            if isinstance(value, KNOWN_TYPES):
                cell_text.append(value)
                if isinstance(value, XLSXCurrency):
                    col_fmt_pairs.append((col, FMT_CURRENCY))
                    row_widths.append(len(value) * text_multiplier)
                elif isinstance(value, basestring):
                    row_widths.append(len(value) * text_multiplier)
                elif value is not None:
                    row_widths.append(len(str(value)) * text_multiplier)
                else:  # Effectively None
                    row_widths.append(None)
            elif isinstance(value, (OpenpyxlImage, PILImage.Image)):
                log.info('embedding image: {}'.format(value))

                if isinstance(value, PILImage.Image):
                    value = self.image(value)
                col_img_pairs.append((col, value))

                # Add empty text data for the cell the image is going into.
                cell_text.append(None)

                row_widths.append(value.drawing.width
                        * prefs['export.xlsx.img.multiplier.width'])
                row_height = max(row_height, value.drawing.height
                        * prefs['export.xlsx.img.multiplier.height'])
            else:
                err = 'Can not handle type \'{}\' in XLSXExporter'.format(
                        type(value).__name__)
                raise TypeError(err)

        if len(cell_text) != len(row_widths):
            err = 'len(cell_text): {} != len(row_widths): {}'.format(
                    len(cell_text), len(row_widths))
            raise ExportError(err)

        self.ws.append(cell_text)

        # wait until after append to get the current row, to cover the
        # ambiguous empty worksheet case.  See below:
        # ws = new worksheet
        # ws.max_row returns 1
        # ws.append(data)
        # ws.max_row returns 1
        row = self.ws.max_row

        self._apply_img(row, col_img_pairs)
        self._apply_row_number_format(row, col_fmt_pairs)
        self._apply_width(row_widths)
        self._apply_row_height(row, row_height)

    def write_hdr(self, iterable):
        """Writes single header row to xlsx file.  All strings are set as bold.

        Args:
            iterable: string or list of strings.
        """
        self.write_row(iterable, multiplier=1.1)
        try:
            # Get row version < 2.4.0
            row = self.ws.rows[self.ws.max_row - 1]
        except:
            # Get row version >= 2.4.0
            row = self.ws[self.ws.max_row]
        for cell in row:
            cell.font = Font(bold=True)

    def write_blank(self, count=1):
        """Write blank lines to csv file.

        Args:
            count (optional): int number of blank lines to write.
        """
        # No openpyxl function to handle this.
        self.ws._current_row += count

    def image(self, img, jpeg=None):
        """Creates a compatible image that may be embedded into a xlsx file.
        If img is an openpyxl image it is returned unchanged.  Other images
        may be resized and/or converted based on preferences.

        Args:
            img: openpyxl image, PIL image, file point to image, string path
                to image.
            jpeg (optional): True to convert image to jpeg.  If not set
                conversion will happen based on preferences.

        Returns:
            openpyxl image.
        """
        if self.img_support is not True:
            return 'image not supported'

        if isinstance(img, OpenpyxlImage):
            return img

        if not isinstance(img, PILImage.Image):
            img = PILImage.open(img)

        # To save space in the xlsx file resize the image using PIL.  Resizing
        # in openpyxl only changes the viewing size but embeds the original
        # sized image.
        size = (prefs['export.image.resize.max_width'],
                prefs['export.image.resize.max_height'])
        # thumbnail will preserve aspect ratio using size as a max width or
        # height and only scale down.
        img.thumbnail(size, resample=PILImage.LANCZOS)

        if jpeg is None:
            jpeg = prefs['export.xlsx.img.jpeg.enable']

        if jpeg is True:
            params = {'quality': prefs['export.xlsx.img.jpeg.quality']}
            img = mp_pil_image_save(img, 'JPEG', params)
            ret = OpenpyxlImageExt(img)
            ret._ext = '.jpeg'
            return ret

        return OpenpyxlImage(img)


def _use_xlsx():
    """Returns true if default exporter is xlsx based on preferences and
    support."""
    return prefs['export.default_csv'] is False and _openpyxl_support() is True


def get_default_export_path():
    """Returns default export path based on preferences and support."""
    name = 'export'
    if _use_xlsx() is True:
        ext = '.xlsx'
    else:
        ext = '.csv'
    return name + ext


def get_exporter_class():
    """Returns exporter class based on preferences and support."""
    if _use_xlsx() is True:
        return XLSXExporter
    else:
        return CSVExporter


def Exporter(path):
    """Returns new exporter object based on preferences and support.

    Args:
        path: string path of export file.  It will be overwritten if exists.
    """
    return get_exporter_class()(path)
